
from binance.client import Client
from datetime import datetime
import pytz
from openpyxl import load_workbook
from datetime import *
import string
import xlsxwriter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from functions import *
from pymexc import spot, futures




settingsWorkbook=load_workbook('settings.xlsx')

#   ------------------------------------    common data
settingsSheet=settingsWorkbook['Common']

fileName=str(settingsSheet["B1"].value)
output_fileName=str(settingsSheet["B2"].value)
timezone=pytz.timezone(str(settingsSheet["B3"].value))

workbook,fileName=checkAndCreateFile(fileName)

#   ------------------------------------    Binance
print("\n ---------------------------- Binance ----------------------------\n")
settingsSheet=settingsWorkbook['Binance']

api_key,api_secret,assetTickers=getExchangeData(settingsSheet)


client = Client(api_key, api_secret)
for assetTicker in list(assetTickers):

    if assetTicker in workbook.sheetnames:
        print(assetTicker+" sheet exists.")
    else:
        print("Warning: "+assetTicker + " sheet doesn't exists, the program creates one.")
        workbook=createSheet(workbook,assetTicker)

    sheet=workbook[assetTicker]
    lastTrade, position=getLastTrade(sheet)
    lastTrade=lastTrade.replace(second=0)

    sheet=handleBinancePair(assetTicker+'USDT',sheet, timezone, client,lastTrade,position)

#   ------------------------------------    MEXC

print("\n ---------------------------- MEXC ----------------------------\n")

settingsSheet=settingsWorkbook['MEXC']

api_key,api_secret,assetTickers=getExchangeData(settingsSheet)

spot_client = spot.HTTP(api_key = api_key, api_secret = api_secret)


for assetTicker in list(assetTickers):

    if assetTicker in workbook.sheetnames:
        print(assetTicker+" sheet exists.")
    else:
        print("\nWarning: "+assetTicker + " sheet doesn't exists, the program creates one.")
        workbook=createSheet(workbook,assetTicker)
    sheet=workbook[assetTicker]

    lastTrade, position=getLastTrade(sheet)
    print(lastTrade)
    lastTrade=lastTrade.replace(second=0)
    # print(assetTicker+'USDT')
    sheet=handleMEXCPair(assetTicker+'USDT',sheet, timezone, spot_client,lastTrade,position)

workbook.save(output_fileName)

