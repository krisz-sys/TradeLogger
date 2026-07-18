from datetime import *

from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import os.path
import xlsxwriter
from pymexc.base import MexcAPIError

# sources
# https://www.codespeedy.com/change-the-number-format-in-openpyxl/

def createSheet(workbook, assetTicker):         # this function creates adn edits the missing wheet in the excel file
    workbook.create_sheet(assetTicker)
    sheet = workbook[assetTicker]
    thin = Side(border_style="thin", color="000000")
    sheet.merge_cells('B3:H3')
    sheet["B3"] = assetTicker
    sheet["B3"].alignment = Alignment(horizontal='center', vertical='center')
    for a in list(map(chr, range(66, 72))):             # border and alignment edit
        sheet[a + "4"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet[a + "4"].alignment = Alignment(horizontal='center', vertical='center')

    # filling the necessary content
    sheet["B4"] = "Date & Time"
    sheet["C4"] = "Ammount"
    sheet["D4"] = "Cost(USD)"
    sheet["E4"] = "Price(USD)"
    sheet["F4"] = "Action"
    sheet["G4"] = "Fee"
    sheet["H4"] = "Platform"
    sheet["I4"] = "Wallet"
    sheet["J4"] = "SUM(TOKEN)"
    sheet["K4"] = "SUM(USD)"
    sheet["L4"] = "Deposit wallet"

    return workbook

def getLastTrade(sheet):                # returns the date and the position of the last trade
    lastTrade = ''
    position = 1
    for i in range(5, 10000):           # the data starts in the row 5
        date = sheet["B" + str(i)].value
        position = i
        if str(date) == 'None':         # in case of an empty row there are no trades
            if lastTrade=='':           # in case of the sheet exists but it is empty than there is no last trade => the program creates one
                lastTrade=datetime.strptime('00/04/12 08:25:03', '%y/%m/%d %H:%M:%S')
            break
        lastTrade = date
    return lastTrade,position



def handleBinancePair(assetPair,sheet, timezone, client,lastTrade,position):            # returns the filled sheet with the new trades
    info = client.get_symbol_info(assetPair)
    if info==None:
        print("Warning: The pair "+assetPair+" isn't exists in the Binance exchange!\n")
        return sheet
    trades = client.get_my_trades(symbol=assetPair)

    data=[]
    for i in enumerate(trades):
        row=[i[1]['time'], i[1]['price'], i[1]['qty'],i[1]['quoteQty'],i[1]['isBuyer'],i[1]['commission'],i[1]['commissionAsset']]
        data.append(row)
    print('\n' + "New positions on binance account(" + assetPair + "): ")
    sheet=handleInsert(data,lastTrade,position,sheet,timezone,["BINANCE(CEX)","-"])
    return sheet


def handleMEXCPair(assetPair,sheet, timezone, client,lastTrade,position):

    trades = []
    isError=False
    try:
        # trades = client.account_trade_list(symbol=assetPair)
        trades = client.call("GET", "api/v3/myTrades", params={}, symbol=assetPair)

        if trades==[]:                                  # the is no trackable trade informations
            print("\nWarning: There is no available trade information about "+assetPair+". Be carrefour that only the last 1 mount of the trade information is trackable by the MEXC API! ")
            isError=True
    except MexcAPIError as err:                         # there is no such pair in this exchange
        if(str(err).find('code=-1121')!=-1):
            print("\nWarning: The pair " + assetPair + " isn't exists!")
            isError = True

    if isError==True:
        return sheet


    data=[]
    for i in enumerate(trades):
        row=[i[1]['time'], i[1]['price'], i[1]['qty'],i[1]['quoteQty'],i[1]['isBuyer'],i[1]['commission'],i[1]['commissionAsset']]
        data.insert(0,row)                                                                  # needs to insert in the beginning of the array becaus the API sends the sequence inverse
    print('\n' + "New positions on MEXC account(" + assetPair + "): ")
    sheet = handleInsert(data, lastTrade, position, sheet, timezone, ["MEXC(CEX)", "-"])
    return sheet



def handleInsert(data,lastTrade,position,sheet,timezone,exchangeSpecific):     # this functions inserts the data to the sheet. Very important the data structur witch is handled by the function witch calls this function
    newTrades = []

    for i in enumerate(data):
        date = i[1][0]
        price = i[1][1]
        quantity = i[1][2]
        invested = i[1][3]
        isBuyer = i[1][4]
        commission=i[1][5]

        querryTrade = datetime.fromtimestamp(round(date / 1000), tz=timezone)
        querryTrade = querryTrade.replace(tzinfo=None)
        querryTrade = querryTrade.replace(second=0)
        action='SELL'
        if str(isBuyer) == "True":
            action='BUY'
        print('date: '+str(querryTrade) + ", price: " + str(price) + ", quantity: " + str(quantity)+', action: '+action+', commission: '+str(commission))
        if querryTrade > lastTrade:
            newTrades.append(i[1])
            sheet["B" + str(position)] = querryTrade
            sheet["B" + str(position)].alignment = Alignment(horizontal='center', vertical='center')

            sheet["C" + str(position)] = float(quantity)
            sheet["C" + str(position)].alignment = Alignment(horizontal='center', vertical='center')

            sheet["D" + str(position)] = float(invested)
            sheet["D" + str(position)].alignment = Alignment(horizontal='center', vertical='center')

            sheet["E" + str(position)] = float(price)
            sheet["E" + str(position)].alignment = Alignment(horizontal='center', vertical='center')

            sheet["F" + str(position)] = action
            sheet["F" + str(position)].alignment = Alignment(horizontal='center', vertical='center')

            sheet["G"+ str(position)] = float(commission)
            sheet["G"+ str(position)].alignment = Alignment(horizontal='center', vertical='center')

            sheet["H" + str(position)] = exchangeSpecific[0]                        # CEX name
            sheet["H" + str(position)].alignment = Alignment(horizontal='center', vertical='center')

            sheet["I" + str(position)] = exchangeSpecific[1]                        # wallet name
            sheet["I" + str(position)].alignment = Alignment(horizontal='center', vertical='center')

            sheet["J" + str(position)] = "= IF($F" + str(position) + " = \"BUY\", C" + str(position) + ", -C" + str(
                position) + ")"
            sheet["J" + str(position)].alignment = Alignment(horizontal='center', vertical='center')

            sheet["K" + str(position)] = "=IF($F" + str(position) + " = \"BUY\", -D" + str(position) + ", D" + str(
                position) + ")"
            sheet["K" + str(position)].alignment = Alignment(horizontal='center', vertical='center')
            position = position + 1
    print('')
    return sheet


def checkAndCreateFile(filename):

    if filename.find('.xlsx')==-1:                       # check if the input file named correctly, and of not correct it
        filename=filename+'.xlsx'
        print("Corrigate the filename of the input name with an '.xlsx'.\n")

    check_file = os.path.isfile(filename)
    if check_file==False:
        print("The "+filename+" file doesn't exists. The program will create a file with this name!\n")
        wb=xlsxwriter.Workbook(filename)
        wb.close()
    workbook=load_workbook(filename)
    return workbook,filename

def getExchangeData(settingsSheet):
    api_key = settingsSheet["B1"].value
    api_secret = settingsSheet["B2"].value

    assetTickers = []
    for a in list(map(chr, range(66, 91))):
        newTicker = str(settingsSheet[a + "3"].value)
        if newTicker == 'None':
            break
        assetTickers.append(newTicker)
    return api_key,api_secret,assetTickers